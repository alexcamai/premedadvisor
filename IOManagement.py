"""
Author: Alex Camai
v 1.0.0
FIOManager.py

Handles user input in the form of data files.

Contains:
    import_courses: Fully imports courses and stores their information in a database.
    write_to_file: Writes the contents of a Schedule to a .xlsx file.
    _load_from_xl: Loads 'proto' course-list (pre_reqs are stored as strings, not Courses) from a .xlsx file.
    _package: Attaches Course prerequisites to each course in a list by searching for them in a database.

"""

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range

import sql

from ScheduleBits import Course, Schedule

_SUBJ = 'A'
_ID = 'B'
_CREDITS = 'C'
_DIFF = 'D'
_PRE = 'E'
_MULTI = 'F'
_DL = 'G'
_TAKEN = 'H'


def import_courses(src: str = "courses.xlsx", db: str = "courses.db"):
    courses, taken = _load_from_xl(src)
    sql.create_db(courses + taken, db)
    return _package(courses, db), _package(taken, db)


def write_to_file(schedule: 'Schedule', *, dest="plan.xlsx"):
    out = Workbook()
    plan = out.active
    plan.title = "Plan"
    # summary = out.create_sheet(title="Summary")

    # Write the column headers
    plan['A1'] = "SEM ->"

    c = 2

    for i in range(0, len(schedule) * 2, 2):
        plan.cell(row=1, column=c + i, value="S" + str((i + 2) // 2))
        plan.cell(row=1, column=c + i + 1, value="CREDITS")

    for index in range(0, len(schedule.semesters)):
        semester_view = iter(schedule.semesters[index].courses.values())
        for row in range(1, len(schedule.semesters[index]) + 1):
            cur = semester_view.__next__()
            plan.cell(row=row + 1, column=c + 2 * index, value=cur.get_course_code())
            plan.cell(row=row + 1, column=c + (2 * index + 1), value=cur.credit_load)
        plan.cell(row=len(schedule.semesters[index]) + 3, column=c + 2 * index, value="TOTAL:")
        plan.cell(row=len(schedule.semesters[index]) + 3, column=c + (2 * index + 1),
                  value=schedule.semesters[index].load)
        plan.cell(row=len(schedule.semesters[index]) + 4, column=c + 2 * index, value="DIFF:")
        plan.cell(row=len(schedule.semesters[index]) + 4, column=c + (2 * index + 1),
                  value=schedule.semesters[index].difficulty)

    out.save(dest)


def _load_from_xl(filename: str):
    current_workbook = load_workbook(filename=filename)
    sheet_ranges = current_workbook.active

    proto_course_list = list()
    proto_taken_list = list()

    r = 2
    # FIXME this method is absolutely disgusting

    while sheet_ranges[_ID + str(r)].value is not None:

        # Required information

        if type(sheet_ranges[_ID + str(r)].value) is not int:
            raise TypeError("Error in cell [{}{}]: type {} was not int".format(_ID, r,
                                                                               type(sheet_ranges
                                                                                    [_ID + str(r)].value)))
        c_id = sheet_ranges[_ID + str(r)].value

        if type(sheet_ranges[_SUBJ + str(r)].value) is not str:
            raise TypeError("Error in cell [{}{}]: type {} was not str".format(_SUBJ, r,
                                                                               type(sheet_ranges
                                                                                    [_SUBJ + str(r)].value)))
        c_subj = sheet_ranges[_SUBJ + str(r)].value.strip()

        if type(sheet_ranges[_CREDITS + str(r)].value) is not int:
            raise TypeError("Error in cell [{}{}]: type {} was not int".format(_ID, r,
                                                                               type(sheet_ranges
                                                                                    [_CREDITS + str(r)].value)))
        c_cred = sheet_ranges[_CREDITS + str(r)].value

        # Options

        if sheet_ranges[_DIFF + str(r)].value is not None and type(sheet_ranges[_DIFF + str(r)].value) is not float:
            raise TypeError("Error in cell [{}{}]: type {} was not float".format(_DIFF, r,
                                                                                 type(sheet_ranges
                                                                                      [_DIFF + str(r)].value)))

        c_diff = sheet_ranges[_DIFF + str(r)].value if sheet_ranges[_DIFF + str(r)].value is not None else 0.5

        pre = list()

        if sheet_ranges[_PRE + str(r)].value is not None:
            pre = sheet_ranges[_PRE + str(r)].value.split(',')

        if sheet_ranges[_MULTI + str(r)].value is not None and type(sheet_ranges[_MULTI + str(r)].value) is not str:
            TypeError("Error in cell [{}{}]: type {} was not str".format(_MULTI, r,
                                                                         type(sheet_ranges
                                                                              [_MULTI + str(r)].value)))

        c_multi = True if sheet_ranges[_MULTI + str(r)].value is not None and \
            sheet_ranges[_MULTI + str(r)].value.lower == 'y' else False

        if sheet_ranges[_DL + str(r)].value is not None and type(sheet_ranges[_DL + str(r)].value) is not int:
            raise TypeError("Error in cell [{}{}]: type {} was not int".format(_DL, r,
                                                                               type(sheet_ranges
                                                                                    [_DL + str(r)].value)))

        c_dl = None if sheet_ranges[_DL + str(r)].value is None else sheet_ranges[_DL + str(r)].value

        if sheet_ranges[_TAKEN + str(r)].value == 'Y':
            proto_taken_list.append(Course(c_id, c_subj, c_cred, diff=c_diff, pre_reqs=pre, multi=c_multi,
                                           deadline=c_dl))
        else:
            proto_course_list.append(Course(c_id, c_subj, c_cred, diff=c_diff, pre_reqs=pre, multi=c_multi,
                                            deadline=c_dl))

        r += 1

    return proto_course_list, proto_taken_list


def _package(proto_course_list: list(), db: str):
    results = list()

    for course in proto_course_list:
        pre = sql.find(course.pre_reqs, db)
        course.pre_reqs = pre
        results.append(course)

    return results
